#!/usr/bin/env python3
"""Fill DPP.json skeleton with actual data from DPP_FIBROTOR_ER15_V2.xlsx."""

import json
import openpyxl

XLSX_PATH = "data/DPP_submodels_specifications_LAG14ER_filled.xlsx"
SKELETON_PATH = "DPP.json"
OUTPUT_PATH = "DPP_filled.json"


def strip_lang(val):
    """Strip trailing '@en', '@de' etc. from xlsx values."""
    if not isinstance(val, str):
        return val
    for suffix in ("@en", "@de", "@fr", "@es"):
        if val.endswith(suffix):
            return val[: -len(suffix)]
    return val


def find_submodel(submodels, id_short):
    for sm in submodels:
        if sm.get("idShort") == id_short:
            return sm
    return None


def find_element(elements, id_short):
    for el in elements:
        if el.get("idShort") == id_short:
            return el
    return None


def make_prop(id_short, value, semantic="https://admin-shell.io/SMT/General/Arbitrary"):
    return {
        "category": "PARAMETER",
        "idShort": id_short,
        "semanticId": {
            "type": "ExternalReference",
            "keys": [{"type": "GlobalReference", "value": semantic}]
        },
        "valueType": "xs:string",
        "value": str(value),
        "modelType": "Property"
    }


# ---------------------------------------------------------------------------
# 1. TechnicalData
# ---------------------------------------------------------------------------

def fill_technical_data(submodels, combined):
    sm = find_submodel(submodels, "TechnicalData")
    if not sm:
        return
    elems = sm["submodelElements"]

    # --- GeneralInformation ---
    gi = find_element(elems, "GeneralInformation")
    if gi:
        for el in gi.get("value", []):
            ids = el.get("idShort", "")
            if ids == "ManufacturerName" and "ManufacturerName" in combined:
                el["value"] = strip_lang(combined["ManufacturerName"])
            elif ids == "CompanyLogo" and "CompanyLogo" in combined:
                el["value"] = combined["CompanyLogo"]
            elif ids == "ManufacturerProductDesignation":
                if "ManufacturerProductDesignation" in combined:
                    el["value"] = [{"language": "en", "text": strip_lang(combined["ManufacturerProductDesignation"])}]
            elif ids == "ManufacturerArticleNumber":
                if "ManufacturerArticleNumber" in combined:
                    el["value"] = combined["ManufacturerArticleNumber"]
            elif ids == "ManufacturerOrderCode":
                if "ManufacturerOrderCode" in combined:
                    el["value"] = combined["ManufacturerOrderCode"]
            elif ids == "ProductImages":
                _fill_product_images(el, combined)

    # --- ProductClassifications ---
    pc = find_element(elems, "ProductClassifications")
    if pc:
        entries = pc.get("value", [])
        if entries:
            for el in entries[0].get("value", []):
                ids = el.get("idShort", "")
                if ids == "ClassificationSystem" and "ClassificationSystem" in combined:
                    el["value"] = combined["ClassificationSystem"]
                elif ids == "ClassificationSystemVersion" and "ClassificationSystemVersion" in combined:
                    el["value"] = combined["ClassificationSystemVersion"]
                elif ids == "ClassificationSystemUrl" and "ClassificationSystemUrl" in combined:
                    el["value"] = combined["ClassificationSystemUrl"]
                elif ids == "ProductClassId" and "ProductClassId" in combined:
                    el["value"] = combined["ProductClassId"]
                elif ids == "ProductClassCodedName" and "ProductClassCodedName" in combined:
                    el["value"] = combined["ProductClassCodedName"]
                elif ids == "ProductClassName" and "ProductClassName" in combined:
                    el["value"] = [{"language": "en", "text": str(combined["ProductClassName"])}]

    # --- TechnicalPropertyAreas: replace arbitrary placeholders ---
    tpa = find_element(elems, "TechnicalPropertyAreas")
    if tpa:
        props = []
        mapping = [
            ("Table_Top_Diameter", None),
            ("Centre_Hole", None),
            ("Weight", None),
            ("Working_Position", None),
            ("Drive_Motor", None),
            ("Drive_Arrangement", None),
            ("Division", None),
            ("Motor_Power", None),
            ("Motor_Voltage", None),
            ("Brake_Voltage", None),
            ("Indexing_Accuracy", None),
            ("Axial_Runout", None),
            ("Concentricity", None),
            ("Plane_Parrallism", None),
            ("Voltage_Motor", None),
            ("Voltage_Brake", None),
            ("Permissible_Transport_Load", None),
            ("Permissible_Addon_Diameter", None),
            ("Permissible_Axial_Loading", None),
            ("Permissible_Radial_Loading", None),
            ("Permissible_Tilting_Moment_Positioned", None),
            ("Permissible_Tilting_Moment_Rotating", None),
            ("Permissible_Tangential_Moment", None),
        ]
        for prop_id, _ in mapping:
            if prop_id in combined:
                props.append(make_prop(prop_id, combined[prop_id]))

        tpa["value"] = [{
            "displayName": [
                {"language": "en", "text": "Technical property area"},
                {"language": "de", "text": "Technischer Merkmalsbereich"}
            ],
            "semanticId": {
                "type": "ExternalReference",
                "keys": [{"type": "GlobalReference",
                          "value": "0173-1#02-ABL358#002/0173-1#01-AHX773#002"}]
            },
            "value": props,
            "modelType": "SubmodelElementCollection"
        }]

    # --- FurtherInformation ---
    fi = find_element(elems, "FurtherInformation")
    if fi:
        for el in fi.get("value", []):
            ids = el.get("idShort", "")
            if ids == "TextStatement" and "TextStatement" in combined:
                el["value"] = [{"language": "en", "text": combined["TextStatement"]}]
            elif ids == "ValidDate" and "ValidDate" in combined:
                el["value"] = combined["ValidDate"]

    # --- SpecificDescriptions ---
    sd = find_element(elems, "SpecificDescriptions")
    if sd and "SpecificDescription" in combined:
        entries = sd.get("value", [])
        if entries:
            for ev in entries[0].get("value", []):
                ids = ev.get("idShort", "")
                if ids == "ArbitraryProperty":
                    ev["value"] = combined["SpecificDescription"]
                elif ids == "ArbitraryMLP":
                    ev["value"] = [{"language": "en", "text": combined["SpecificDescription"]}]


def _fill_product_images(pi_el, data):
    entries = pi_el.get("value", [])
    if not entries:
        return
    inner = entries[0].get("value", [])
    for ev in inner:
        ids = ev.get("idShort", "")
        if ids == "ImageFile" and "ProductImages" in data:
            ev["value"] = data["ProductImages"]
        elif ids == "ImageNote" and "ImageNote" in data:
            ev["value"] = [{"language": "en", "text": data["ImageNote"]}]


# ---------------------------------------------------------------------------
# 2. Nameplate
# ---------------------------------------------------------------------------

def fill_nameplate(submodels, data):
    sm = find_submodel(submodels, "Nameplate")
    if not sm:
        return
    for el in sm.get("submodelElements", []):
        ids = el.get("idShort", "")
        if ids == "ManufacturerName":
            el["value"] = [{"language": "en", "text": strip_lang(data.get("ManufacturerName", ""))}]
        elif ids == "ManufacturerProductDesignation":
            el["value"] = [{"language": "en", "text": strip_lang(data.get("ManufacturerProductDesignation", ""))}]
        elif ids == "AddressInformation":
            _fill_address(el, data)
        elif ids == "ManufacturerProductRoot" and "ManufacturerProductRoot" in data:
            el["value"] = [{"language": "en", "text": data["ManufacturerProductRoot"]}]
        elif ids == "ManufacturerProductFamily" and "ManufacturerProductFamily" in data:
            el["value"] = [{"language": "en", "text": data["ManufacturerProductFamily"]}]
        elif ids == "ManufacturerProductType" and "ManufacturerProductType" in data:
            el["value"] = data["ManufacturerProductType"]
        elif ids == "OrderCodeOfManufacturer" and "OrderCodeOfManufacturer" in data:
            el["value"] = data["OrderCodeOfManufacturer"]
        elif ids == "ProductArticleNumberOfManufacturer" and "ProductArticleNumberOfManufacturer" in data:
            el["value"] = data["ProductArticleNumberOfManufacturer"]
        elif ids == "CountryOfOrigin" and "CountryOfOrigin" in data:
            el["value"] = data["CountryOfOrigin"]
        elif ids == "SerialNumber" and "SerialNumber" in data:
            el["value"] = data["SerialNumber"]
        elif ids == "YearOfConstruction" and "YearOfConstruction" in data:
            el["value"] = data["YearOfConstruction"]
        elif ids == "DateOfManufacture" and "DateOfManufacture" in data:
            el["value"] = data["DateOfManufacture"]
        elif ids == "HardwareVersion" and "HardwareVersion" in data:
            el["value"] = data["HardwareVersion"]
        elif ids == "FirmwareVersion" and "FirmwareVersion" in data:
            el["value"] = data["FirmwareVersion"]
        elif ids == "SoftwareVersion" and "SoftwareVersion" in data:
            el["value"] = data["SoftwareVersion"]
        elif ids == "UniqueFacilityIdentifier" and "UniqueFacilityIdentifier" in data:
            el["value"] = data["UniqueFacilityIdentifier"]


def _fill_address(addr_el, data):
    addr_str = data.get("AddressInformation", "")
    street = zipcode = city = country = ""
    if addr_str:
        parts = [p.strip() for p in addr_str.split(",")]
        street = parts[0] if len(parts) > 0 else ""
        rest = parts[1].strip() if len(parts) > 1 else ""
        if rest:
            tokens = rest.split(" ", 1)
            zipcode = tokens[0]
            city_country = tokens[1] if len(tokens) > 1 else ""
            if city_country:
                cc = city_country.rsplit(" ", 1)
                city = cc[0] if len(cc) > 1 else city_country
                country = cc[1] if len(cc) > 1 else ""

    smc = {"idShort": "AddressInformation", "value": [], "modelType": "SubmodelElementCollection"}
    for short, val in [("Street", street), ("Zipcode", zipcode), ("CityTown", city), ("NationalCode", country)]:
        if val:
            smc["value"].append({
                "idShort": short,
                "value": [{"language": "en", "text": val}],
                "modelType": "MultiLanguageProperty"
            })
    if smc["value"]:
        addr_el["value"] = [smc]


# ---------------------------------------------------------------------------
# 3. CarbonFootprint
# ---------------------------------------------------------------------------

def fill_carbon_footprint(submodels, data):
    sm = find_submodel(submodels, "CarbonFootprint")
    if not sm:
        return
    pcf_list = find_element(sm["submodelElements"], "ProductCarbonFootprints")
    if not pcf_list:
        return
    entries = pcf_list.get("value", [])
    if not entries:
        return
    for fv in entries[0].get("value", []):
        ids = fv.get("idShort", "")
        if ids == "PcfCO2eq" and "PcfCO2eq" in data:
            fv["value"] = str(data["PcfCO2eq"])
        elif ids == "ReferenceImpactUnitForCalculation" and "ReferenceImpactUnitForCalculation" in data:
            fv["value"] = str(data["ReferenceImpactUnitForCalculation"])
        elif ids == "QuantityOfMeasureForCalculation" and "QuantityOfMeasureForCalculation" in data:
            fv["value"] = str(data["QuantityOfMeasureForCalculation"])
        elif ids == "PublicationDate" and "PublicationDate" in data:
            fv["value"] = str(data["PublicationDate"])


# ---------------------------------------------------------------------------
# 4. HandoverDocumentation
# ---------------------------------------------------------------------------

def read_handover_docs(ws):
    """Parse the Handover Documentation xlsx sheet into a list of 2 document dicts.
    Each dict has keys: Title, Subtitle, Description, KeyWords, StatusSetDate,
    StatusValue, OrganizationShortName, OrganizationOfficialName, DigitalFiles."""
    docs = []
    current_doc = {}
    in_doc_version = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        a_val = str(row[0].value).strip() if row[0].value else ""
        g_val = row[6].value if len(row) > 6 and row[6].value else None

        # Start of a DocumentVersion block
        if a_val.startswith("2.8") and "DocumentVersion" in a_val:
            in_doc_version = True
            continue

        if not in_doc_version:
            continue

        # End of a DocumentVersion block — any new section header that is NOT
        # "2.8 ..." or "2.8.1 ..." means we left the block
        if a_val.startswith("2.") and "Submodel Elements" in a_val:
            if not a_val.startswith("2.8"):
                in_doc_version = False
                if current_doc:
                    docs.append(current_doc)
                    current_doc = {}
                continue

        # We are inside a "2.8 Submodel Elements of 'DocumentVersion'" block
        if a_val == "Title" and g_val is not None:
            # Start of a new document version — save previous if it had a Title
            if "Title" in current_doc:
                docs.append(current_doc)
            current_doc = {"Title": strip_lang(str(g_val))}
        elif a_val == "Subtitle" and g_val is not None:
            current_doc["Subtitle"] = strip_lang(str(g_val))
        elif a_val == "Description" and g_val is not None:
            current_doc["Description"] = strip_lang(str(g_val))
        elif a_val == "KeyWords" and g_val is not None:
            current_doc["KeyWords"] = strip_lang(str(g_val))
        elif a_val == "StatusSetDate" and g_val is not None:
            current_doc["StatusSetDate"] = str(g_val)[:10]
        elif a_val == "StatusValue" and g_val is not None:
            current_doc["StatusValue"] = str(g_val)
        elif a_val == "OrganizationShortName" and g_val is not None:
            current_doc["OrganizationShortName"] = strip_lang(str(g_val))
        elif a_val == "OrganizationOfficialName" and g_val is not None:
            current_doc["OrganizationOfficialName"] = strip_lang(str(g_val))
        elif a_val == "Version" and g_val is not None:
            current_doc["Version"] = str(g_val)
        elif a_val == "File" and g_val is not None:
            current_doc["DigitalFiles"] = str(g_val)

    if current_doc:
        docs.append(current_doc)

    return docs


def make_doc_version(doc):
    title = doc.get("Title", "")
    subtitle = doc.get("Subtitle", "")
    description = doc.get("Description", "")
    keywords = doc.get("KeyWords", "")
    status_date = doc.get("StatusSetDate", "")
    status_val = doc.get("StatusValue", "Released")
    org_short = doc.get("OrganizationShortName", "")
    org_official = doc.get("OrganizationOfficialName", "")
    version = doc.get("Version", "1.0")
    digital_files = doc.get("DigitalFiles", "")

    value_elements = [
        {"idShort": "Language", "displayName": [{"language": "en", "text": "Language"}],
         "semanticId": {"type": "ExternalReference",
                        "keys": [{"type": "GlobalReference", "value": "0173-1#02-AAN468#008"}]},
         "typeValueListElement": "Property", "valueTypeListElement": "xs:string",
         "value": [{"displayName": [{"language": "en", "text": "en (English)"}],
                    "semanticId": {"type": "ExternalReference",
                                   "keys": [{"type": "GlobalReference", "value": "0173-1#02-AAN468#008"}]},
                    "valueType": "xs:string", "value": "en",
                    "valueId": {"type": "ExternalReference",
                                "keys": [{"type": "GlobalReference", "value": "0173-1#07-AAS045#003"}]},
                    "modelType": "Property"}],
         "modelType": "SubmodelElementList"},
        {"category": "PARAMETER", "idShort": "Version",
         "displayName": [{"language": "en", "text": "Document version"}],
         "semanticId": {"type": "ExternalReference",
                        "keys": [{"type": "GlobalReference", "value": "0173-1#02-AAP003#005"}]},
         "valueType": "xs:string", "value": version, "modelType": "Property"},
        {"category": "PARAMETER", "idShort": "Title",
         "displayName": [{"language": "en", "text": "Document title"}],
         "semanticId": {"type": "ExternalReference",
                        "keys": [{"type": "GlobalReference", "value": "0173-1#02-ABG940#003"}]},
         "value": [{"language": "en", "text": title}], "modelType": "MultiLanguageProperty"},
    ]
    if subtitle:
        value_elements.append(
            {"category": "PARAMETER", "idShort": "Subtitle",
             "displayName": [{"language": "en", "text": "Subtitle"}],
             "semanticId": {"type": "ExternalReference",
                            "keys": [{"type": "GlobalReference", "value": "0173-1#02-ABH998#003"}]},
             "value": [{"language": "en", "text": subtitle}], "modelType": "MultiLanguageProperty"})
    value_elements.append(
        {"category": "PARAMETER", "idShort": "Description",
         "displayName": [{"language": "en", "text": "Document description"}],
         "semanticId": {"type": "ExternalReference",
                        "keys": [{"type": "GlobalReference", "value": "0173-1#02-AAN466#004"}]},
         "value": [{"language": "en", "text": description}], "modelType": "MultiLanguageProperty"})
    if keywords:
        value_elements.append(
            {"category": "PARAMETER", "idShort": "KeyWords",
             "displayName": [{"language": "en", "text": "Keywords"}],
             "semanticId": {"type": "ExternalReference",
                            "keys": [{"type": "GlobalReference", "value": "0173-1#02-ABH999#003"}]},
             "value": [{"language": "en", "text": keywords}], "modelType": "MultiLanguageProperty"})
    value_elements += [
        {"category": "PARAMETER", "idShort": "StatusSetDate",
         "displayName": [{"language": "en", "text": "Document status set date"}],
         "semanticId": {"type": "ExternalReference",
                        "keys": [{"type": "GlobalReference", "value": "0173-1#02-ABI000#003"}]},
         "valueType": "xs:date", "value": status_date, "modelType": "Property"},
        {"category": "PARAMETER", "idShort": "StatusValue",
         "displayName": [{"language": "en", "text": "Document status"}],
         "semanticId": {"type": "ExternalReference",
                        "keys": [{"type": "GlobalReference", "value": "0173-1#02-ABI001#003"}]},
         "valueType": "xs:string", "value": status_val, "modelType": "Property"},
        {"category": "PARAMETER", "idShort": "OrganizationShortName",
         "displayName": [{"language": "en", "text": "Organization short name"}],
         "semanticId": {"type": "ExternalReference",
                        "keys": [{"type": "GlobalReference",
                                  "value": "https://api.eclass-cdp.com/0173-1-02-ABI002-003"}]},
         "valueType": "xs:string", "value": org_short, "modelType": "Property"},
        {"category": "PARAMETER", "idShort": "OrganizationOfficialName",
         "displayName": [{"language": "en", "text": "Organization official name"}],
         "semanticId": {"type": "ExternalReference",
                        "keys": [{"type": "GlobalReference", "value": "0173-1#02-ABI004#003"}]},
         "valueType": "xs:string", "value": org_official, "modelType": "Property"},
        {"idShort": "RefersToEntities", "displayName": [{"language": "en", "text": "Reference to other documents"}],
         "semanticId": {"type": "ExternalReference",
                        "keys": [{"type": "GlobalReference", "value": "0173-1#02-ABK288#002"}]},
         "typeValueListElement": "ReferenceElement",
         "value": [{"category": "PARAMETER",
                    "semanticId": {"type": "ExternalReference",
                                   "keys": [{"type": "GlobalReference", "value": "0173-1#02-ABK288#002"}]},
                    "modelType": "ReferenceElement"}],
         "modelType": "SubmodelElementList"},
        {"idShort": "BasedOnReferences", "displayName": [{"language": "en", "text": "Based on other documents"}],
         "semanticId": {"type": "ExternalReference",
                        "keys": [{"type": "GlobalReference", "value": "0173-1#02-ABK289#002"}]},
         "typeValueListElement": "ReferenceElement",
         "value": [{"category": "PARAMETER",
                    "semanticId": {"type": "ExternalReference",
                                   "keys": [{"type": "GlobalReference", "value": "0173-1#02-ABK289#002"}]},
                    "modelType": "ReferenceElement"}],
         "modelType": "SubmodelElementList"},
        {"idShort": "TranslationOfEntities",
         "displayName": [{"language": "en", "text": "Translation of other documents"}],
         "semanticId": {"type": "ExternalReference",
                        "keys": [{"type": "GlobalReference", "value": "0173-1#02-ABK290#002"}]},
         "typeValueListElement": "ReferenceElement",
         "value": [{"displayName": [{"language": "en", "text": "Translation of documents"}],
                    "semanticId": {"type": "ExternalReference",
                                   "keys": [{"type": "GlobalReference", "value": "0173-1#02-ABK290#002"}]},
                    "modelType": "ReferenceElement"}],
         "modelType": "SubmodelElementList"},
        {"idShort": "DigitalFiles", "displayName": [{"language": "en", "text": "Digital files"}],
         "semanticId": {"type": "ExternalReference",
                        "keys": [{"type": "GlobalReference", "value": "0173-1#02-ABK126#002"}]},
         "typeValueListElement": "File",
         "value": [{"displayName": [{"language": "en", "text": title[:64]},
                                    {"language": "de", "text": title[:64]}],
                    "value": digital_files,
                    "semanticId": {"type": "ExternalReference",
                                   "keys": [{"type": "GlobalReference", "value": "0173-1#02-ABK126#002"}]},
                    "contentType": "application/pdf", "modelType": "File"}],
         "modelType": "SubmodelElementList"},
        {"category": "PARAMETER", "idShort": "PreviewFile",
         "displayName": [{"language": "en", "text": "Preview file"}],
         "semanticId": {"type": "ExternalReference",
                        "keys": [{"type": "GlobalReference", "value": "0173-1#02-ABK127#002"}]},
         "contentType": "image/jpeg", "modelType": "File"}
    ]

    return {
        "displayName": [
            {"language": "en", "text": "Document version"},
            {"language": "de", "text": "Document version"}
        ],
        "semanticId": {
            "type": "ExternalReference",
            "keys": [{"type": "GlobalReference",
                      "value": "0173-1#02-ABI503#003/0173-1#01-AHF582#003"}]
        },
        "value": value_elements,
        "modelType": "SubmodelElementCollection"
    }


def fill_handover_documentation(submodels, docs_data):
    sm = find_submodel(submodels, "HandoverDocumentation")
    if not sm:
        return
    docs_el = find_element(sm["submodelElements"], "Documents")
    if not docs_el:
        return
    doc_entries = docs_el.get("value", [])
    if not doc_entries:
        return
    first_entry = doc_entries[0]
    dv = find_element(first_entry.get("value", []), "DocumentVersions")
    if not dv:
        return
    dv["value"] = [make_doc_version(d) for d in docs_data]


# ---------------------------------------------------------------------------
# 5. MaintenanceInstructions
# ---------------------------------------------------------------------------

def fill_maintenance_instructions(submodels, data):
    sm = find_submodel(submodels, "MaintenanceInstructions")
    if not sm:
        return
    elems = sm["submodelElements"]

    mfa = find_element(elems, "MaintenanceFreeAsset")
    if mfa and "MaintenanceFreeAsset" in data:
        mfa["value"] = data["MaintenanceFreeAsset"]

    interval_el = find_element(elems, "MaintenanceInstructionsForSpecificInterval__00__")
    if not interval_el:
        return
    interval_values = interval_el.get("value", [])

    bmi = find_element(interval_values, "BasicMaintenanceInformation")
    if not bmi:
        # Add BasicMaintenanceInformation if not in skeleton
        bmi = {
            "idShort": "BasicMaintenanceInformation",
            "semanticId": {
                "type": "ExternalReference",
                "keys": [{"type": "GlobalReference",
                          "value": "https://adminshell.io/idta/maintenanceinstructions/basicmaintenanceinformation/1/0"}]
            },
            "value": [],
            "modelType": "SubmodelElementCollection"
        }
        interval_values.insert(0, bmi)

    for bmv in bmi.get("value", []):
        ids = bmv.get("idShort", "")
        if ids == "MaintenanceID" and "MaintenanceID" in data:
            bmv["value"] = data["MaintenanceID"]
        elif ids == "NameOfMaintenance" and "NameOfMaintenance" in data:
            bmv["value"] = [{"language": "en", "text": data["NameOfMaintenance"]}]
        elif ids == "SourceOfMaintenanceInstructions" and "SourceOfMaintenanceInstructions" in data:
            bmv["value"] = [{"language": "en", "text": data["SourceOfMaintenanceInstructions"]}]
        elif ids == "SafetyRegulationsToBeObserved" and "SafetyRegulationsToBeObserved" in data:
            bmv["value"] = [{"language": "en", "text": data["SafetyRegulationsToBeObserved"]}]
        elif ids == "IntervalSpecification":
            for isv in bmv.get("value", []):
                isv_ids = isv.get("idShort", "")
                if isv_ids == "MaintenanceIntervalValue" and "MaintenanceIntervalValue" in data:
                    isv["value"] = data["MaintenanceIntervalValue"]
                elif isv_ids == "MaintenanceIntervalUnit" and "MaintenanceIntervalUnit" in data:
                    isv["value"] = data["MaintenanceIntervalUnit"]
        elif ids == "ContactForMaintenanceAuthorization":
            for cv in bmv.get("value", []):
                cv_ids = cv.get("idShort", "")
                if cv_ids == "Company" and "ContactCompany" in data:
                    cv["value"] = [{"language": "en", "text": data["ContactCompany"]}]
                elif cv_ids == "Street" and "ContactStreet" in data:
                    cv["value"] = [{"language": "de", "text": data["ContactStreet"]}]
                elif cv_ids == "Zipcode" and "ContactZipcode" in data:
                    cv["value"] = [{"language": "de", "text": data["ContactZipcode"]}]
                elif cv_ids == "CityTown" and "ContactCity" in data:
                    cv["value"] = [{"language": "de", "text": data["ContactCity"]}]
                elif cv_ids == "NationalCode" and "ContactCountry" in data:
                    cv["value"] = [{"language": "de", "text": data["ContactCountry"]}]

    # Now populate the BMI fields
    bmi_fields = bmi.get("value", [])
    if not bmi_fields:
        # Build the full BMI structure from xlsx data
        bmi["value"] = [
            {"idShort": "MaintenanceID",
             "semanticId": {"type": "ExternalReference",
                            "keys": [{"type": "GlobalReference",
                                      "value": "https://adminshell.io/idta/maintenanceinstructions/maintenanceid/1/0"}]},
             "valueType": "xs:string",
             "value": data.get("MaintenanceID", ""),
             "modelType": "Property"},
            {"idShort": "NameOfMaintenance",
             "semanticId": {"type": "ExternalReference",
                            "keys": [{"type": "GlobalReference",
                                      "value": "https://adminshell.io/idta/maintenanceinstructions/nameofmaintenance/1/0"}]},
             "value": [{"language": "en", "text": data.get("NameOfMaintenance", "")}],
             "modelType": "MultiLanguageProperty"},
            {"idShort": "SourceOfMaintenanceInstructions",
             "semanticId": {"type": "ExternalReference",
                            "keys": [{"type": "GlobalReference",
                                      "value": "https://adminshell.io/idta/maintenanceinstructions/sourceofmaintenanceinstructions/1/0"}]},
             "value": [{"language": "en", "text": data.get("SourceOfMaintenanceInstructions", "")}],
             "modelType": "MultiLanguageProperty"},
            {"idShort": "SafetyRegulationsToBeObserved",
             "semanticId": {"type": "ExternalReference",
                            "keys": [{"type": "GlobalReference",
                                      "value": "https://adminshell.io/idta/maintenanceinstructions/safetyregulationstobeobserved/1/0"}]},
             "value": [{"language": "en", "text": data.get("SafetyRegulationsToBeObserved", "")}],
             "modelType": "MultiLanguageProperty"},
            {"idShort": "ContactForMaintenanceAuthorization",
             "semanticId": {"type": "ExternalReference",
                            "keys": [{"type": "GlobalReference",
                                      "value": "https://adminshell.io/zvei/nameplate/1/0/ContactInformations/ContactInformation"}]},
             "value": [
                 {"idShort": "Company",
                  "value": [{"language": "en", "text": data.get("ContactCompany", "")}],
                  "modelType": "MultiLanguageProperty"},
                 {"idShort": "Street",
                  "value": [{"language": "de", "text": data.get("ContactStreet", "")}],
                  "modelType": "MultiLanguageProperty"},
                 {"idShort": "Zipcode",
                  "value": [{"language": "de", "text": data.get("ContactZipcode", "")}],
                  "modelType": "MultiLanguageProperty"},
                 {"idShort": "CityTown",
                  "value": [{"language": "de", "text": data.get("ContactCity", "")}],
                  "modelType": "MultiLanguageProperty"},
                 {"idShort": "NationalCode",
                  "value": [{"language": "de", "text": data.get("ContactCountry", "")}],
                  "modelType": "MultiLanguageProperty"},
             ],
             "modelType": "SubmodelElementCollection"},
        ]


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    with open(SKELETON_PATH, "r", encoding="utf-8") as f:
        dpp = json.load(f)

    submodels = dpp["submodels"]
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    # --- Sheet1 summary ---
    ws1 = wb["Sheet1"]
    summary = {}
    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row):
        a = row[0].value
        b = row[1].value if len(row) > 1 else None
        if a and b:
            summary[str(a).strip()] = b

    # --- Technical Data ---
    ws_tech = wb["Technical Data"]
    tech = {}
    for row in ws_tech.iter_rows(min_row=1, max_row=ws_tech.max_row):
        a = str(row[0].value).strip() if row[0].value else ""
        g = row[6].value if len(row) > 6 and row[6].value else None
        if a and g:
            tech[a] = g

    # Merge Sheet1 into tech for fields not in Technical Data sheet
    sheet1_map = {
        "Centre hole": "Centre_Hole",
        "Weight": "Weight",
        "Working position": "Working_Position",
        "Motor voltage": "Motor_Voltage",
        "Brake voltage": "Brake_Voltage",
        "Motor power": "Motor_Power",
        "ER indexing accuracy": "Indexing_Accuracy",
        "Horizontal transport load": "Permissible_Transport_Load",
        "Horizontal add-on diameter": "Permissible_Addon_Diameter",
        "Horizontal axial loading": "Permissible_Axial_Loading",
        "Horizontal radial loading": "Permissible_Radial_Loading",
        "Horizontal tilting moment, positioned": "Permissible_Tilting_Moment_Positioned",
        "Horizontal tilting moment, rotating": "Permissible_Tilting_Moment_Rotating",
        "Horizontal tangential moment, positioned": "Permissible_Tangential_Moment",
    }
    combined = dict(tech)
    for src_key, tgt_key in sheet1_map.items():
        if src_key in summary and tgt_key not in combined:
            combined[tgt_key] = summary[src_key]

    # --- Fill submodels ---
    print("Filling TechnicalData...")
    fill_technical_data(submodels, combined)

    print("Filling Nameplate...")
    np_data = {
        "ManufacturerName": summary.get("Manufacturer", "Fibro Rundtische GmbH"),
        "ManufacturerProductDesignation": "FIBROTOR ER.15 indexing table",
        "AddressInformation": summary.get("Address", ""),
        "ManufacturerProductRoot": "rotary indexing table",
        "ManufacturerProductFamily": "FIBROTOR ER.15",
        "ManufacturerProductType": "ER.15",
        "OrderCodeOfManufacturer": summary.get("Product designation / order code", ""),
        "ProductArticleNumberOfManufacturer": summary.get("Manufacturer article number", ""),
        "CountryOfOrigin": summary.get("Country of origin", "DE"),
    }
    fill_nameplate(submodels, np_data)

    # print("Filling CarbonFootprint...")
    # fill_carbon_footprint(submodels, {
    #     "PcfCO2eq": "0",
    #     "ReferenceImpactUnitForCalculation": "1 piece",
    #     "QuantityOfMeasureForCalculation": "1",
    #     "PublicationDate": "2026-06-11T12:00:00Z",
    # })

    print("Filling HandoverDocumentation...")
    ws_docs = wb["Handover Documentation"]
    docs_data = read_handover_docs(ws_docs)
    print(f"  Found {len(docs_data)} document versions")
    for d in docs_data:
        print(f"    - {d.get('Title', '?')}")
    fill_handover_documentation(submodels, docs_data)

    print("Filling MaintenanceInstructions...")
    fill_maintenance_instructions(submodels, {
        "MaintenanceFreeAsset": "false",
        "MaintenanceID": "MI-ER15-PROTOTYPE-001",
        "NameOfMaintenance": "Periodic inspection and maintenance for FIBROTOR ER.15",
        "SourceOfMaintenanceInstructions": "Fibro operating manual MBA-DE-03-01 / MBA-EN-03-01",
        "SafetyRegulationsToBeObserved": "Before maintenance, follow the operating manual and machine safety procedures; disconnect energy sources and secure against restart where applicable.",
        "ContactCompany": "Fibro Rundtische GmbH",
        "ContactStreet": "Weidachstraße 41-43",
        "ContactZipcode": "74189",
        "ContactCity": "Weinsberg",
        "ContactCountry": "DE",
    })

    # --- Write output ---
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(dpp, f, indent=2, ensure_ascii=False)

    print(f"\nDone! Output: {OUTPUT_PATH}")
    print(f"Size: {len(json.dumps(dpp, indent=2, ensure_ascii=False)):,} chars")


if __name__ == "__main__":
    main()
